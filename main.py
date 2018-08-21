# coding:utf-8
import openpyxl

x = 2

#打开文件
w1 = openpyxl.load_workbook('1.xlsx')
w2 = openpyxl.load_workbook('3.xlsx')

# 新建record的xlsx文件
w3 = openpyxl.Workbook()
sheet3 = w3.active
sheet3.title = 'record'

#获取sheet名称
a = w1.sheetnames
b = w2.sheetnames
#只有一页sheet，取第一页分析
sheet1 = w1.get_sheet_by_name(a[0])
sheet2 = w2.get_sheet_by_name(b[0])
#获取最大行数
max1 = sheet1.max_row
max2 = sheet2.max_row
#获取最大列数
m1 = sheet1.max_column
m2 = sheet2.max_column
#遍历，判断是否是待投资名单上的学校，若是则写入新表
for i in range(2,max2 + 1):
    for j in range(2,max1 + 1):
        if sheet1.cell(row=j,column=1).value == sheet2.cell(row=i,column=1).value:
            for k in range(1,m1+1):
                sheet3.cell(row=x,column=k).value = sheet1.cell(row=j,column=k).value
            x += 1
#保存
w3.save('record.xlsx')





